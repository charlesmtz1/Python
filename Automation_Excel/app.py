import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook("transactions.xlsx")

sheet = wb["Sheet1"]

cell = sheet.cell(1, 1)


for row in range(2, sheet.max_row + 1):
    #cell = sheet.cell(row, 3)
    #precio_corregido = cell.value * 0.9
    #celda_corregida = sheet.cell(row, 4)
    #celda_corregida.value = precio_corregido
    sheet.cell(row, 5).value = sheet.cell(row, 3).value * 0.9
    #print (precio_corregido)

values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=5, max_col=5)

chart = BarChart()

chart.add_data(values)
sheet.add_chart(chart, "F2")

wb.save("transactions2.xlsx")